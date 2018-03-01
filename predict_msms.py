# -*- coding: utf-8 -*-
"""
Created on Tue Nov 15 10:49:37 2016

@author: nune558
"""

from subprocess import PIPE, Popen
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from time import time

#%% Functions
def _cmdline(command, mode):
    process = Popen(
        args=command,
        stdout=PIPE,
        shell=True,
        cwd = r'C:\Users\nune558\Google Drive\Jamie Nunez PNNL\EPA\CFM ID\\' + mode
    )
    return process.communicate()[0]

def specify_energy(full_result, energy):
    full_result = full_result.replace('\r', '')
    lines = full_result.split('\n')
    start = lines.index('energy' + str(energy)) + 1
    if energy < 2:
        stop = lines.index('energy' + str(energy+1))
    else:
        stop = len(lines)
    output = ''
    for piece in lines[start:stop]:
        output += piece + '\n'
    return output.strip()

# Mode can be positive ('+' or anything starting with 'p') or negative ('-' or 
# anything starting with 'n')
def predict(inchi, mode, energy=None):
    if mode == '+' or mode.lower()[0] == 'p':
        m = 'Positive Mode'
    elif mode == '-' or mode.lower()[0] == 'n':
        m = 'Negative Mode'
    else:
        print 'Error. Unknown mode \"' + mode + '\".'
        return None
    result = _cmdline('cfm-predict.exe ' + inchi, m)
    if energy is None:
        return result.strip()
    else:
        return specify_energy(result, energy)
        
def _plot_spectra(spectra):
    spectra = spectra.strip().split('\n')
    x = []
    y = []
    for piece in spectra:
        xy = piece.split(' ')
        x.append(float(xy[0]))
        y.append(float(xy[1][:-1]))
    markerline, stemlines, baseline = plt.stem(x, y, '-', lw=3)
    plt.axis([0 , 185, 0, m])
    plt.show()

# Assumes the spectra is the combination of all three energies. Input energy=0-2 for a specific
# energy or energy < 0 if the input already has a specified energy
def plot_spectra(spectra, e=None):
    if spectra is not None:
        if e is None:
            print '\n\nEnergy = 10V'
            _plot_spectra(specify_energy(spectra, 0))
            print '\n\nEnergy = 20V'
            _plot_spectra(specify_energy(spectra, 1))
            print '\n\nEnergy = 40V'
            _plot_spectra(specify_energy(spectra, 2))
        elif e >= 0 and e <= 2:
            _plot_spectra(specify_energy(spectra, e))
        else:
            _plot_spectra(spectra)
    else:
        print 'Error. Can not plot \'NoneType\' spectra.'

#%% Example
#m = 50
#inchi = 'InChI=1S/C6H12O6/c7-1-3(9)5(11)6(12)4(10)2-8/h1,3-6,8-12H,2H2/t3-,4+,5+,6+/m0/s1'; m=50 # D-Glucose
##inchi = 'InChI=1S/C8H10N4O2/c1-10-4-9-6-5(10)7(13)12(3)8(14)11(6)2/h4H,1-3H3'; m=30 # Caffeine
#print plot_spectra(predict(inchi, '+'))
#print '\n'
#print plot_spectra(predict(inchi, '-'))

#%% DSSTox Predictions
path = r'C:\Users\nune558\Google Drive\Jamie Nunez PNNL\SFMetab\HMDB\\'
name = 'HMDB_ISICLE'
wb = load_workbook(path + name + '.xlsx')
sheet = wb.worksheets[0]
maxx = sheet.max_row + 1
t = time()
for i in range(2, maxx):
    inchi = str(sheet['B'+str(i)].value)
    sheet['J'+str(i)] = predict(inchi, '+')
    sheet['K'+str(i)] = predict(inchi, '-')
    if i % 100 == 0:
        print 'Time/cpd: %.2f' % ((time()-t) / i)
        wb.save(path + name + '_MSMS.xlsx')

wb.save(path + name + '_MSMS.xlsx')
print 'Total time: ' + '%.2f' % (time()-t) + ' s'